# -*- coding: utf-8 -*-
"""Convert FC_RCA_Holiday_Master_Production.xlsx into the compact JSON the engine reads.

Run once (and again whenever the master is reissued):

    cd backend
    python load_holiday_master.py "..\\..\\Final Files\\Final Files\\FC_RCA_Holiday_Master_Production.xlsx"

WHY A BUILD STEP RATHER THAN READING THE XLSX DIRECTLY
--------------------------------------------------------
Three reasons, in order of importance:

1. `.gitignore` excludes `*.xlsx`, so the master itself can never be committed. A JSON
   extract can, which means the holiday calendar travels with the code instead of being a
   file someone has to remember to copy onto every machine.
2. Parsing 12,197 rows of xlsx on every RCA request would add seconds to each call for
   data that changes once a year.
3. openpyxl would become a runtime dependency of the backend. The engine is deliberately
   dependency-light; keeping the parser in a build script leaves the runtime on the
   standard library.

WHAT IS KEPT
------------
Only what the engine can actually reason with: country, fiscal week, canonical name, type,
and the per-holiday impact window. The evidence, validation and change-log sheets stay in
the workbook -- they are provenance for humans, not inputs.

Rows with `Active_Flag` false are dropped. `Requires_Review` is carried through so a name
that was inferred rather than confirmed can be reported as such instead of being quoted
with false authority.
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

OUT = Path(__file__).resolve().parent / "wfm" / "context_repository" / "holiday_master.json"


def build(xlsx_path):
    import openpyxl                      # build-time only, never imported at runtime

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["01_Holiday_Master"]
    it = ws.iter_rows(min_row=4, values_only=True)
    hdr = next(it)
    idx = {h: i for i, h in enumerate(hdr) if h}

    def col(row, name, default=None):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else default

    by_country_week = defaultdict(list)
    kept = dropped = 0
    for row in it:
        if not row or col(row, "Fiscal_Week") is None:
            continue
        if not col(row, "Active_Flag"):
            dropped += 1
            continue
        try:
            fw = int(col(row, "Fiscal_Week"))
        except (TypeError, ValueError):
            dropped += 1
            continue
        country = str(col(row, "Country") or "").strip().lower()
        name = (col(row, "Canonical_Name_Final") or col(row, "Name") or "holiday")
        entry = {
            "name": str(name).strip(),
            "type": str(col(row, "Type") or "").strip(),
            "date": str(col(row, "Date_ISO") or col(row, "Date") or ""),
            "before": int(col(row, "Impact_Days_Before") or 0),
            "after": int(col(row, "Impact_Days_After") or 0),
            "group": str(col(row, "Aggregate_Group") or "").strip(),
            # A name inferred from an adjacent week is not the same as a confirmed one.
            "needs_review": bool(col(row, "Requires_Review")),
        }
        # The master holds one row per (holiday, country, type); several sources can list
        # the same holiday. Deduplicate on name so a week does not report "Christmas Day,
        # Christmas Day, Christmas Day".
        bucket = by_country_week[f"{country}|{fw}"]
        if not any(e["name"].lower() == entry["name"].lower() for e in bucket):
            bucket.append(entry)
        kept += 1

    # Aggregate groups (AMER_GROUP etc.) let a queue whose Country is an aggregate value
    # still resolve to something, per sheet 07_Aggregate_Mapping.
    groups = defaultdict(set)
    for key, entries in by_country_week.items():
        country = key.split("|")[0]
        for e in entries:
            if e.get("group"):
                groups[e["group"]].add(country)

    fiscal = {}
    if "06_Fiscal_Calendar" in wb.sheetnames:
        fws = wb["06_Fiscal_Calendar"]
        fit = fws.iter_rows(min_row=4, values_only=True)
        fhdr = next(fit)
        fidx = {h: i for i, h in enumerate(fhdr) if h}
        for r in fit:
            if not r or r[fidx["Fiscal_Week"]] is None:
                continue
            fiscal[str(int(r[fidx["Fiscal_Week"]]))] = {
                "start": str(r[fidx["Week_Start"]]), "end": str(r[fidx["Week_End"]]),
                "quarter": r[fidx["Quarter"]], "month": r[fidx["Month"]]}

    payload = {
        "source": Path(xlsx_path).name,
        "active_rows": kept,
        "inactive_dropped": dropped,
        "country_weeks": len(by_country_week),
        "holidays": {k: v for k, v in by_country_week.items()},
        "aggregate_groups": {g: sorted(c) for g, c in groups.items()},
        "fiscal_calendar": fiscal,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, separators=(",", ":"), ensure_ascii=False),
                   encoding="utf-8")
    return payload


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else (
        Path(__file__).resolve().parents[2] / "Final Files" / "Final Files"
        / "FC_RCA_Holiday_Master_Production.xlsx")
    p = build(src)
    print(f"  source          : {p['source']}")
    print(f"  active rows     : {p['active_rows']:,}  (dropped {p['inactive_dropped']:,} inactive)")
    print(f"  country-weeks   : {p['country_weeks']:,}")
    print(f"  aggregate groups: {len(p['aggregate_groups'])}")
    print(f"  fiscal weeks    : {len(p['fiscal_calendar']):,}")
    print(f"  written to      : {OUT}")
