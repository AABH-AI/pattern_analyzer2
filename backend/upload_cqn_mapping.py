"""
Load the CQN / Forecast-Name mapping workbook into SQL Server.

    python upload_cqn_mapping.py --dry-run       # parse + show CREATE TABLE + coverage, NO database
    python upload_cqn_mapping.py                 # load dbo.CQN_Mapping (Sheet1: flat + dimensions)
    python upload_cqn_mapping.py --pairs         # ALSO load dbo.CQN_Forecast_Pair (Sheet3 Data Pair)
    python upload_cqn_mapping.py --verify-sheets # prove Sheet1 and Sheet3 carry the same mapping
    python upload_cqn_mapping.py --coverage      # after loading: report unmapped queues

WHICH SHEET? (measured, not assumed)
------------------------------------
The workbook's TAB ORDER is `Sheet2, Sheet3, Sheet1` — so the *third tab* is the one **named
`Sheet1`**, which is confusing when someone says "the 3rd sheet".

| Tab | Name | Shape | Verdict |
|-----|------|-------|---------|
| 1 | `Sheet2` | pivot: Row Labels + Count of Forecast_Name | a count, not a mapping — not loaded |
| 2 | `Sheet3` | pivot: Combined_Queue_Name + Forecast_Name, **191 of 523 CQN cells blank** | the same mapping in PIVOT form (blank = repeat of the group above) |
| 3 | `Sheet1` | flat: Region · SubRegion · Channel · Offering · Forecast_Name · Combined_Queue_Name · DB_OSP | **authoritative** |

Verified: forward-fill `Sheet3`'s blank CQN cells (as its pivot layout implies) and the two sheets
agree on **442/442 Forecast_Names — 100%**, same 331 CQNs, same 69 one-to-many fan-outs. They are
the same data. `Sheet1` is preferred only because it needs no blank-inference and carries the five
extra dimension columns required to join to `Input_To_ML`.

Read `Sheet3` naively — without forward-filling — and 167 of 442 names come out mapped to blank.
`--verify-sheets` exists so that claim can be re-checked rather than trusted.

WHY THIS MATTERS -- it settles a real ambiguity
----------------------------------------------
`rca_console.html:1648-1653` records the client's CQN as
`Forecast_name + Region + SubRegion + Country + Channel`, i.e. **channel is part of the key**.
The business RCA prompt instead asks for demand migration *between channels within the same CQN*,
which is only possible if channel is NOT in the key. Those two cannot both be true.

This workbook decides it: of 331 distinct Combined_Queue_Names, **35 span more than one channel**
(e.g. `EMEA English ProSupp Client (Multi-Site)` covers Case, Chat, Email and Voice across 9
forecast names; `ProSupport BRZ Voice` covers Chat, Email and Voice). So the true CQN does span
channels, the prompt's requirement is coherent, and the console's channel-inclusive definition is
a different concept — a locality key, not the Combined Queue.

Once this table exists, the WFM engine resolves channel siblings from the REAL CQN instead of the
locality proxy, and reports `is_cqn_proxy: false`.

The path comes from `--excel`, or `cqn_mapping_path` in config.json.
"""
import argparse
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
DEFAULT_TABLE = "dbo.CQN_Mapping"
PAIR_TABLE = "dbo.CQN_Forecast_Pair"
SHEET = "Sheet1"
PAIR_SHEET = "Sheet3"

# Every column is a short dimension string; no measures in this workbook.
COLUMNS = ("Region", "SubRegion", "Channel", "Offering",
           "Forecast_Name", "Combined_Queue_Name", "DB_OSP")
WIDTH = {"Region": 100, "SubRegion": 100, "Channel": 60, "Offering": 60,
         "Forecast_Name": 255, "Combined_Queue_Name": 255, "DB_OSP": 30}


def load_config(path):
    p = Path(path)
    if not p.is_absolute():
        p = HERE / p
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def read_mapping(xlsx_path):
    """Sheet1 -> list of dicts, blank rows dropped, values trimmed."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        sys.exit(f"Sheet '{SHEET}' not found. Sheets present: {wb.sheetnames}")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    if not rows:
        sys.exit("Sheet1 is empty.")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        sys.exit(f"Sheet1 is missing expected column(s): {missing}\nFound: {header}")

    out = []
    for raw in rows[1:]:
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        rec = dict(zip(header, raw))
        clean = {}
        for c in COLUMNS:
            v = rec.get(c)
            clean[c] = None if v is None or str(v).strip() == "" else str(v).strip()
        # A row with no Forecast_Name cannot be joined to anything.
        if not clean["Forecast_Name"]:
            continue
        out.append(clean)
    return out


PAIR_COLUMNS = ("Forecast_Name", "Combined_Queue_Name", "Source_Sheet", "CQN_Was_Blank_In_Pivot")


def read_pairs(xlsx_path, sheet=PAIR_SHEET):
    """Read the Data Pair sheet (Combined_Queue_Name, Forecast_Name).

    It is laid out as an Excel PIVOT: the CQN is written once per group and left blank on the
    rows beneath it. Those blanks are forward-filled, which is the only reading consistent with
    the data -- 191 of 523 CQN cells are empty, and filling them makes this sheet agree with
    Sheet1 on 442/442 names. The `(blank)` Forecast_Name row is a pivot artefact and is dropped.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        sys.exit("Sheet '%s' not found. Sheets present: %s" % (sheet, wb.sheetnames))
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    header_at = None
    for i, r in enumerate(rows):
        if r and any(str(v).strip() == "Combined_Queue_Name" for v in r if v is not None):
            header_at = i
            break
    if header_at is None:
        sys.exit("Sheet '%s' has no Combined_Queue_Name header." % sheet)
    header = [str(h).strip() if h is not None else "" for h in rows[header_at]]

    out, carried, blanks = [], None, 0
    for raw in rows[header_at + 1:]:
        if not raw or not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        rec = dict(zip(header, raw))
        cqn = rec.get("Combined_Queue_Name")
        fc = rec.get("Forecast_Name")
        cqn = str(cqn).strip() if cqn is not None and str(cqn).strip() else None
        fc = str(fc).strip() if fc is not None and str(fc).strip() else None
        if cqn:
            carried = cqn
        else:
            blanks += 1
        if not fc or fc == "(blank)":
            continue
        out.append({"Forecast_Name": fc,
                    "Combined_Queue_Name": cqn or carried,
                    "Source_Sheet": sheet,
                    "CQN_Was_Blank_In_Pivot": "N" if cqn else "Y"})
    print("  [%s] pair rows %d; forward-filled %d blank CQN cell(s)" % (sheet, len(out), blanks))
    return out


def verify_sheets(xlsx_path):
    """Prove Sheet1 and Sheet3 carry the same mapping. Returns 0 if identical."""
    from collections import defaultdict
    flat = read_mapping(xlsx_path)
    pairs = read_pairs(xlsx_path)
    m1, m3 = defaultdict(set), defaultdict(set)
    for r in flat:
        m1[r["Forecast_Name"]].add(r["Combined_Queue_Name"])
    for r in pairs:
        m3[r["Forecast_Name"]].add(r["Combined_Queue_Name"])
    both = set(m1) & set(m3)
    disagree = sorted(f for f in both if m1[f] != m3[f])
    print("")
    print("--- sheet cross-check ---")
    print("  %s      : %d names / %d CQNs" % (SHEET, len(m1), len({c for v in m1.values() for c in v})))
    print("  %s (ffill): %d names / %d CQNs" % (PAIR_SHEET, len(m3), len({c for v in m3.values() for c in v})))
    print("  overlapping names : %d" % len(both))
    print("  only in %s   : %s" % (SHEET, sorted(set(m1) - set(m3))[:5]))
    print("  only in %s   : %s" % (PAIR_SHEET, sorted(set(m3) - set(m1))[:5]))
    print("  DISAGREEMENTS     : %d" % len(disagree))
    for f in disagree[:10]:
        print("    %-44s %s=%s  %s=%s" % (f[:44], SHEET, sorted(m1[f]), PAIR_SHEET, sorted(m3[f])))
    ok = (not disagree) and set(m1) == set(m3)
    print("  => %s" % ("IDENTICAL mapping in both sheets" if ok else "THE SHEETS DO NOT AGREE"))
    return 0 if ok else 1


def build_create_pairs(table):
    cols = ",\n    ".join([
        "[Forecast_Name] NVARCHAR(255) NULL",
        "[Combined_Queue_Name] NVARCHAR(255) NULL",
        "[Source_Sheet] NVARCHAR(30) NULL",
        "[CQN_Was_Blank_In_Pivot] NVARCHAR(1) NULL"])
    return ("IF OBJECT_ID('%s','U') IS NOT NULL DROP TABLE %s;\n"
            "CREATE TABLE %s (\n    %s\n);\n"
            "CREATE INDEX IX_CQN_Pair_Forecast_Name ON %s ([Forecast_Name]);\n"
            "CREATE INDEX IX_CQN_Pair_CQN ON %s ([Combined_Queue_Name]);"
            % (table, table, table, cols, table, table))


def build_create(table):
    cols = ",\n    ".join(f"[{c}] NVARCHAR({WIDTH[c]}) NULL" for c in COLUMNS)
    return (f"IF OBJECT_ID('{table}','U') IS NOT NULL DROP TABLE {table};\n"
            f"CREATE TABLE {table} (\n    {cols}\n);\n"
            # Joined on Forecast_Name on every investigation; indexed so it stays cheap.
            f"CREATE INDEX IX_CQN_Mapping_Forecast_Name ON {table} ([Forecast_Name]);\n"
            f"CREATE INDEX IX_CQN_Mapping_CQN ON {table} ([Combined_Queue_Name]);")


def summarise(records):
    from collections import defaultdict
    fc = {r["Forecast_Name"] for r in records}
    cqn = {r["Combined_Queue_Name"] for r in records if r["Combined_Queue_Name"]}
    by_cqn_channel = defaultdict(set)
    for r in records:
        if r["Combined_Queue_Name"]:
            by_cqn_channel[r["Combined_Queue_Name"]].add(r["Channel"])
    multi = {c: sorted(v) for c, v in by_cqn_channel.items() if len(v) > 1}
    print(f"  rows                              : {len(records)}")
    print(f"  distinct Forecast_Name            : {len(fc)}")
    print(f"  distinct Combined_Queue_Name      : {len(cqn)}")
    print(f"  CQNs spanning >1 channel          : {len(multi)}  "
          f"({100*len(multi)/max(len(cqn),1):.1f}%)  <- channel migration is real")
    for c, ch in sorted(multi.items(), key=lambda kv: -len(kv[1]))[:5]:
        print(f"      {c[:46]:46s} {ch}")
    dupes = len(records) - len({(r["Forecast_Name"], r["Combined_Queue_Name"]) for r in records})
    if dupes:
        print(f"  duplicate (Forecast_Name, CQN) pairs: {dupes}")
    many = [f for f in fc if len({r["Combined_Queue_Name"] for r in records
                                  if r["Forecast_Name"] == f}) > 1]
    if many:
        print(f"  NOTE: {len(many)} Forecast_Name(s) map to MORE THAN ONE CQN, so a join can fan out:")
        for f in sorted(many)[:6]:
            cs = sorted({r['Combined_Queue_Name'] for r in records if r['Forecast_Name'] == f})
            print(f"      {f[:44]:44s} -> {cs}")
    return many


def coverage(cur, data_table, map_table):
    cur.execute(f"SELECT COUNT(DISTINCT Forecast_name) FROM {data_table}")
    total = cur.fetchone()[0]
    cur.execute(f"""SELECT COUNT(*) FROM (
                      SELECT DISTINCT d.Forecast_name FROM {data_table} d
                       WHERE NOT EXISTS (SELECT 1 FROM {map_table} m
                                          WHERE m.Forecast_Name = d.Forecast_name)) x""")
    unmapped = cur.fetchone()[0]
    print(f"  queues in {data_table}: {total}")
    print(f"  mapped                 : {total - unmapped}  ({100*(total-unmapped)/max(total,1):.1f}%)")
    print(f"  UNMAPPED               : {unmapped}")
    if unmapped:
        cur.execute(f"""SELECT DISTINCT d.Forecast_name FROM {data_table} d
                         WHERE NOT EXISTS (SELECT 1 FROM {map_table} m
                                            WHERE m.Forecast_Name = d.Forecast_name)
                         ORDER BY d.Forecast_name""")
        for r in cur.fetchall():
            print(f"      - {r[0]}")
    return unmapped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", help="Path to the CQN mapping workbook (.xlsx)")
    ap.add_argument("--config", default="config.json")
    ap.add_argument("--table", default=DEFAULT_TABLE)
    ap.add_argument("--dry-run", action="store_true", help="Parse + print schema/summary; no database.")
    ap.add_argument("--coverage", action="store_true", help="Report unmapped queues and exit.")
    ap.add_argument("--pairs", action="store_true",
                    help="Also load the Sheet3 Data Pair list into dbo.CQN_Forecast_Pair.")
    ap.add_argument("--pair-table", default=PAIR_TABLE)
    ap.add_argument("--verify-sheets", action="store_true",
                    help="Prove Sheet1 and Sheet3 carry the same mapping, then exit.")
    args = ap.parse_args()

    cfg = load_config(args.config)
    data_table = (cfg.get("sql") or {}).get("table", "dbo.Input_To_ML")

    if args.verify_sheets:
        xlsx0 = args.excel or cfg.get("cqn_mapping_path")
        if not xlsx0:
            sys.exit("No workbook path. Pass --excel PATH or set cqn_mapping_path in config.json.")
        sys.exit(verify_sheets(xlsx0))

    if args.coverage:
        import pyodbc  # noqa: F401
        from sql_backend import connect
        conn = connect(cfg)
        try:
            print(f"--- coverage: {data_table} vs {args.table} ---")
            sys.exit(0 if coverage(conn.cursor(), data_table, args.table) == 0 else 1)
        finally:
            conn.close()

    xlsx = args.excel or cfg.get("cqn_mapping_path")
    if not xlsx:
        sys.exit("No workbook path. Pass --excel PATH or set cqn_mapping_path in config.json.")
    if not Path(xlsx).exists():
        sys.exit(f"Workbook not found: {xlsx}")

    print(f"--- reading {xlsx} (sheet {SHEET}) ---")
    records = read_mapping(xlsx)
    summarise(records)
    print("\n--- CREATE TABLE ---")
    print(build_create(args.table))

    if args.dry_run:
        print("\n[dry-run] nothing written to the database.")
        return

    from sql_backend import connect
    conn = connect(cfg)
    try:
        cur = conn.cursor()
        for stmt in build_create(args.table).split(";\n"):
            if stmt.strip():
                cur.execute(stmt)
        conn.commit()
        cur.fast_executemany = True
        placeholders = ",".join("?" * len(COLUMNS))
        cur.executemany(
            f"INSERT INTO {args.table} ([{'],['.join(COLUMNS)}]) VALUES ({placeholders})",
            [tuple(r[c] for c in COLUMNS) for r in records])
        conn.commit()
        cur.execute(f"SELECT COUNT(*) FROM {args.table}")
        print(f"\nloaded {cur.fetchone()[0]} rows into {args.table}")

        if args.pairs:
            pair_rows = read_pairs(xlsx)
            for stmt in build_create_pairs(args.pair_table).split(";\n"):
                if stmt.strip():
                    cur.execute(stmt)
            conn.commit()
            cur.executemany(
                "INSERT INTO %s ([Forecast_Name],[Combined_Queue_Name],[Source_Sheet],"
                "[CQN_Was_Blank_In_Pivot]) VALUES (?,?,?,?)" % args.pair_table,
                [tuple(r[c] for c in PAIR_COLUMNS) for r in pair_rows])
            conn.commit()
            cur.execute("SELECT COUNT(*) FROM %s" % args.pair_table)
            total_pairs = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM %s WHERE CQN_Was_Blank_In_Pivot='Y'" % args.pair_table)
            print("loaded %d rows into %s (%d forward-filled from pivot blanks)"
                  % (total_pairs, args.pair_table, cur.fetchone()[0]))
        print(f"\n--- coverage: {data_table} vs {args.table} ---")
        coverage(cur, data_table, args.table)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
