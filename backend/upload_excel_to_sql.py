# -*- coding: utf-8 -*-
"""
Load the weekly demand Excel file into a SQL Server table.

Usage:
    python upload_excel_to_sql.py                 # uses backend/config.json
    python upload_excel_to_sql.py --dry-run       # parse + show CREATE TABLE + row count, NO database
    python upload_excel_to_sql.py --excel PATH --config PATH

Column typing is driven by the known schema of Input_To_ML; unknown columns
default to NVARCHAR(255). config.json holds the connection details (see
config.example.json). config.json is gitignored so credentials are never committed.
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent

# --- column typing for the known Input_To_ML schema ---
NUMERIC = {
    "Actual_Offered", "Actual_Handled", "fcst_offered", "fcst_handled",
    "Planned_ASU", "Actual_ASU", "Final_Units", "Final_Y5", "Final_Y4",
    "Final_Y3", "Final_Y2", "Final_Y1", "Final_upp_units", "Holiday_Count",
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
}
INT_COLS = {"Fiscal_Week"}
DATE_COLS = {"Week_Ending"}
BATCH = 5000


def sql_type(col):
    if col in DATE_COLS:
        return "DATE"
    if col in INT_COLS:
        return "BIGINT"
    if col in NUMERIC:
        return "FLOAT"
    return "NVARCHAR(255)"


def coerce(col, v):
    if v is None or v == "":
        return None
    if col in DATE_COLS:
        return v  # openpyxl returns a datetime for date cells; pyodbc handles it
    if col in INT_COLS:
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None
    if col in NUMERIC:
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    return str(v)


def load_config(path):
    p = Path(path)
    if not p.exists():
        sys.exit(f"Config not found: {p}\nCopy config.example.json to config.json and fill it in.")
    return json.loads(p.read_text(encoding="utf-8"))


def read_excel(path):
    print(f"Reading {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(it) if h is not None]
    rows = list(it)
    wb.close()
    print(f"  {len(headers)} columns, {len(rows):,} data rows")
    return headers, rows


def build_create(table, headers):
    cols = ",\n    ".join(f"[{h}] {sql_type(h)}" for h in headers)
    return f"IF OBJECT_ID('{table}','U') IS NOT NULL DROP TABLE {table};\nCREATE TABLE {table} (\n    {cols}\n);"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default=str(HERE / "config.json"))
    ap.add_argument("--excel", default=None)
    ap.add_argument("--dry-run", action="store_true", help="Parse + print schema only; no database.")
    ap.add_argument("--schema-only", action="store_true", help="Create the table (schema) in SQL only; do not insert any rows.")
    args = ap.parse_args()

    excel_path = args.excel
    table = "dbo.Input_To_ML"
    cfg = {}
    if not args.dry_run or Path(args.config).exists():
        try:
            cfg = load_config(args.config)
            table = cfg.get("sql", {}).get("table", table)
            excel_path = excel_path or cfg.get("excel_path")
        except SystemExit:
            if not args.dry_run:
                raise
    if not excel_path:
        sys.exit("No Excel path. Pass --excel PATH or set excel_path in config.json.")

    headers, rows = read_excel(excel_path)
    create_sql = build_create(table, headers)
    print("\n--- CREATE TABLE ---")
    print(create_sql)

    if args.dry_run:
        # show a coerced sample so typing can be eyeballed without a DB
        print("\n--- sample coerced row (first data row) ---")
        if rows:
            sample = {h: coerce(h, rows[0][i] if i < len(rows[0]) else None) for i, h in enumerate(headers)}
            for k, v in sample.items():
                print(f"  {k:18s} = {v!r}")
        print(f"\nDRY RUN OK — {len(rows):,} rows ready to load into {table}. No database was touched.")
        return

    import pyodbc  # imported here so --dry-run works without the driver
    conn = connect(cfg)
    cur = conn.cursor()
    print(f"\nCreating table {table} ...")
    for stmt in create_sql.split(";"):
        if stmt.strip():
            cur.execute(stmt)
    conn.commit()

    if args.schema_only:
        bare = table.split(".")[-1].strip("[]")
        cur.execute(
            "SELECT COLUMN_NAME, DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_NAME = ? ORDER BY ORDINAL_POSITION", bare)
        print("\nTable created. Columns now in SQL:")
        for cn, dt in cur.fetchall():
            print(f"  {cn}: {dt}")
        conn.close()
        print(f"\nSCHEMA ONLY — {table} created with {len(headers)} columns, 0 rows. "
              f"Re-run without --schema-only to load the {len(rows):,} data rows.")
        return

    placeholders = ",".join("?" * len(headers))
    col_list = ",".join(f"[{h}]" for h in headers)
    insert_sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"
    cur.fast_executemany = True

    print(f"Inserting {len(rows):,} rows (batches of {BATCH}) ...")
    done = 0
    for start in range(0, len(rows), BATCH):
        chunk = rows[start:start + BATCH]
        data = [[coerce(headers[i], r[i] if i < len(r) else None) for i in range(len(headers))] for r in chunk]
        cur.executemany(insert_sql, data)
        conn.commit()
        done += len(chunk)
        print(f"  {done:,}/{len(rows):,}")
    conn.close()
    print(f"\nDone — loaded {done:,} rows into {table}.")


def connect(cfg):
    import pyodbc
    c = cfg["sql"]
    driver = c.get("driver", "ODBC Driver 17 for SQL Server")
    if str(c.get("auth", "sql")).lower() == "windows":
        conn_str = f"DRIVER={{{driver}}};SERVER={c['server']};DATABASE={c['database']};Trusted_Connection=yes;"
    else:
        conn_str = (f"DRIVER={{{driver}}};SERVER={c['server']};DATABASE={c['database']};"
                    f"UID={c.get('username','')};PWD={c.get('password','')};")
    if c.get("encrypt") is not None:
        conn_str += f"Encrypt={'yes' if c['encrypt'] else 'no'};"
    if c.get("trust_server_certificate"):
        conn_str += "TrustServerCertificate=yes;"
    return pyodbc.connect(conn_str, timeout=int(c.get("timeout", 30)))


if __name__ == "__main__":
    main()
